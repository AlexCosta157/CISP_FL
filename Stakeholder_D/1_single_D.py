
#import flwr as fl
import tensorflow as tf
from openpyxl import Workbook
import pandas as pd


IMAGE_SHAPE = (224, 224)
TRAINING_DATA_DIR = 'trainD/'
VALID_DATA_DIR = 'valid/'

EPOCHS = 200 #1
BATCH_SIZE = 32

datagen = tf.keras.preprocessing.image.ImageDataGenerator(
    rescale=1./255
)

train_generator = datagen.flow_from_directory(
    TRAINING_DATA_DIR,
    shuffle=True,
    target_size=IMAGE_SHAPE,
)

valid_generator = datagen.flow_from_directory(
    VALID_DATA_DIR,
    shuffle=False,
    target_size=IMAGE_SHAPE,
)

base_model = tf.keras.applications.MobileNetV2(input_shape=(224, 224, 3),
                                               include_top=False,
                                               weights='imagenet')

base_model.trainable = False
global_average_layer = tf.keras.layers.GlobalAveragePooling2D()
prediction_layer = tf.keras.layers.Dense(2, activation='softmax')

model = tf.keras.Sequential([
  base_model,
  global_average_layer,
  prediction_layer
])

base_learning_rate = 0.0001
model.compile(optimizer=tf.keras.optimizers.RMSprop(learning_rate=base_learning_rate),
              loss=tf.keras.losses.CategoricalCrossentropy(from_logits=False),
              metrics=['accuracy', 'AUC', 'Recall', 'Precision'])

history = model.fit(train_generator,
                    steps_per_epoch=train_generator.samples // BATCH_SIZE,
                    epochs=EPOCHS,
                    batch_size=BATCH_SIZE,
                    validation_data=valid_generator)

pd.DataFrame.from_dict(history.history).to_csv('3_singleDHistory_test.csv', index=False)

loss, accuracy, auc, recall, precision = model.evaluate(valid_generator)
wb = Workbook()
ws = wb.active
ws['B2'] = "Model"
ws['B3'] = "5SingleD"
ws['C2'] = "Loss"
ws['C3'] = loss
ws['D2'] = "accuracy"
ws['D3'] = accuracy
ws['E2'] = "AUC"
ws['E3'] = auc
ws['F2'] = "Recall"
ws['F3'] = recall
ws['G2'] = "Precision"
ws['G3'] = precision
wb.save("3_SingleD_results_test.xlsx")
