import argparse
import os

import tensorflow as tf
from tensorflow_privacy.privacy.optimizers.dp_optimizer_keras_vectorized import (
    VectorizedDPKerasSGDOptimizer,
)
from tensorflow_privacy.privacy.analysis.rdp_accountant import compute_rdp
from tensorflow_privacy.privacy.analysis.rdp_accountant import get_privacy_spent
from keras_preprocessing.image import ImageDataGenerator
import pandas as pd
from openpyxl import Workbook

import flwr as fl


DPSGD: bool = True
LEARNING_RATE = 0.15
NOISE_MULTIPLIER = 1.1
L2_NORM_CLIP = 1.0
BATCH_SIZE = 32
LOCAL_EPOCHS = 5
MICROBATCHES = 32
MODEL_DIR = None

IMAGE_SHAPE = (224, 224)
TRAINING_DATA_DIR = 'trainD/'
VALID_DATA_DIR = 'valid/'

# Make TensorFlow logs less verbose
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"

# global for tracking privacy
PRIVACY_LOSS = 0


def compute_epsilon(
    epochs: int, num_train_examples: int, batch_size: int, noise_multiplier: float
) -> float:
    """Computes epsilon value for given hyperparameters.
    Based on
    github.com/tensorflow/privacy/blob/master/tutorials/mnist_dpsgd_tutorial_keras.py
    """
    if noise_multiplier == 0.0:
        return float("inf")
    steps = epochs * num_train_examples // batch_size
    orders = [1 + x / 10.0 for x in range(1, 100)] + list(range(12, 64))
    sampling_probability = batch_size / num_train_examples
    rdp = compute_rdp(
        q=sampling_probability,
        noise_multiplier=noise_multiplier,
        steps=steps,
        orders=orders,
    )
    # Delta is set to approximate 1 / (number of training points).
    return get_privacy_spent(orders, rdp, target_delta=1 / num_train_examples)[0]



# Define Flower client
class MnistClient(fl.client.NumPyClient):
    def __init__(self, model, train_generator, valid_generator):
        self.model = model
        self.train = train_generator
        self.test = valid_generator
        self.batch_size = BATCH_SIZE
        self.local_epochs = LOCAL_EPOCHS
        self.dpsgd = DPSGD

        if DPSGD:
            self.noise_multiplier = NOISE_MULTIPLIER
            if BATCH_SIZE % MICROBATCHES != 0:
                raise ValueError(
                    "Number of microbatches should divide evenly batch_size"
                )
            optimizer = VectorizedDPKerasSGDOptimizer(
                l2_norm_clip=L2_NORM_CLIP,
                noise_multiplier=NOISE_MULTIPLIER,
                num_microbatches=MICROBATCHES,
                learning_rate=LEARNING_RATE,
            )
            # Compute vector of per-example loss rather than its mean over a minibatch.
            loss = tf.keras.losses.CategoricalCrossentropy(
                from_logits=False, reduction=tf.losses.Reduction.NONE
            )
        else:
            optimizer = tf.keras.optimizers.SGD(learning_rate=LEARNING_RATE)
            loss = tf.keras.losses.CategoricalCrossentropy(from_logits=False)

        # Compile model with Keras
        model.compile(optimizer=optimizer,
                      loss=loss,
                      metrics=['accuracy', 'AUC', 'Recall', 'Precision'])

    def get_parameters(self):
        """Get parameters of the local model."""
        raise Exception("Not implemented (server-side parameter initialization)")

    def fit(self, parameters, config):
        """Train parameters on the locally held training set."""
        # Update local model parameters
        global PRIVACY_LOSS
        if self.dpsgd:
            privacy_spent = compute_epsilon(
                self.local_epochs,
                len(self.train.filenames),
                self.batch_size,
                self.noise_multiplier,
            )
            print('For delta=1e-5, the current epsilon is: %.2f' % privacy_spent)
            PRIVACY_LOSS += privacy_spent

        self.model.set_weights(parameters)
        # Train the model
        history = self.model.fit(self.train,
                                 epochs=self.local_epochs,
                                 batch_size=self.batch_size,
                                 validation_data=self.test
                                 )
        pd.DataFrame.from_dict(history.history).to_csv('DPclientD_FEDAVG_History.csv', index=False)

        results = {
            "loss": history.history["loss"][0],
            "accuracy": history.history["accuracy"][0],
            #"val_loss": history.history["val_loss"][0],
            #"val_accuracy": history.history["val_accuracy"][0],
        }

        return self.model.get_weights(), len(self.train.filenames), results

    def evaluate(self, parameters, config):
        """Evaluate parameters on the locally held test set."""

        # Update local model with global parameters
        self.model.set_weights(parameters)

        # Evaluate global model parameters on the local test data and return results
        loss, accuracy, auc, recall, precision = self.model.evaluate(self.test)

        wb = Workbook()
        ws = wb.active
        ws['B2'] = "Model"
        ws['B3'] = "DP_FedD_FEDAVG"
        ws['C2'] = "Loss"
        ws['C3'] = loss
        ws['D2'] = "accuracy"
        ws['D3'] = accuracy
        ws['E2'] = "AUC"
        ws['E3'] = auc
        ws['F2'] = "Recall"
        ws['F3'] = recall
        ws['G2'] = "Precision"
        ws['G3'] = precision
        wb.save("DP_FedD_FEDAVG.xlsx")

        num_examples_test = len(self.test.filenames)
        return loss, num_examples_test, {"accuracy": accuracy}


def main() -> None:
    # Load Keras model
    base_model = tf.keras.applications.MobileNetV2(input_shape=(224, 224, 3),
                                                   include_top=False,
                                                   weights='imagenet')

    base_model.trainable = False
    global_average_layer = tf.keras.layers.GlobalAveragePooling2D()
    prediction_layer = tf.keras.layers.Dense(2, activation='softmax')

    model = tf.keras.Sequential([
        base_model,
        global_average_layer,
        prediction_layer
    ])

    # Load a subset of MNIST to simulate the local data partition
    datagen = ImageDataGenerator(
        rescale=1. / 255
    )

    train_generator = datagen.flow_from_directory(
        TRAINING_DATA_DIR,
        shuffle=True,
        target_size=IMAGE_SHAPE,
    )

    valid_generator = datagen.flow_from_directory(
        VALID_DATA_DIR,
        shuffle=False,
        target_size=IMAGE_SHAPE,
    )

    # drop samples to form exact batches for dpsgd
    # this is necessary since dpsgd is sensitive to uneven batches
    # due to microbatching
    if DPSGD and len(train_generator.filenames) % BATCH_SIZE != 0:
        drop_num = len(train_generator.filenames) % BATCH_SIZE
        print("len is", len(train_generator.filenames), "it has more", drop_num, "than expected")

    # Start Flower client
    client = MnistClient(model, train_generator, valid_generator)
    fl.client.start_numpy_client("192.168.1.127:8080", client=client)
    if DPSGD:
        print("Privacy Loss: ", PRIVACY_LOSS)


if __name__ == "__main__":
    main()
