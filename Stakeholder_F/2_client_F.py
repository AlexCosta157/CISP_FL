
import flwr as fl
import tensorflow as tf
from openpyxl import Workbook
import pandas as pd


IMAGE_SHAPE = (224, 224)
TRAINING_DATA_DIR = 'trainE224/'
# VALID_DATA_DIR = 'trainE416/'

EPOCHS = 5  # 1
BATCH_SIZE = 32

datagen = tf.keras.preprocessing.image.ImageDataGenerator(
    rescale=1./255,
    validation_split=0.2
)

train_generator = datagen.flow_from_directory(
    TRAINING_DATA_DIR,
    shuffle=True,
    target_size=IMAGE_SHAPE,
    subset='training'
)

valid_generator = datagen.flow_from_directory(
    TRAINING_DATA_DIR,
    shuffle=False,
    target_size=IMAGE_SHAPE,
    subset='validation'
)

base_model = tf.keras.applications.MobileNetV2(input_shape=(224, 224, 3),
                                               include_top=False,
                                               weights='imagenet')

base_model.trainable = False
global_average_layer = tf.keras.layers.GlobalAveragePooling2D()
prediction_layer = tf.keras.layers.Dense(2, activation='softmax')

model = tf.keras.Sequential([
  base_model,
  global_average_layer,
  prediction_layer
])

base_learning_rate = 0.0001
model.compile(optimizer=tf.keras.optimizers.RMSprop(learning_rate=base_learning_rate),
              loss=tf.keras.losses.CategoricalCrossentropy(from_logits=False),
              metrics=['accuracy', 'AUC', 'Recall', 'Precision'])

# Define Flower client
class CifarClient(fl.client.NumPyClient):
    def get_parameters(self):
        return model.get_weights()

    def fit(self, parameters, config):
        model.set_weights(parameters)
        history = model.fit(train_generator,
                            steps_per_epoch=train_generator.samples // BATCH_SIZE,
                            epochs=EPOCHS,
                            batch_size=BATCH_SIZE,
                            validation_data=valid_generator
                            )
        pd.DataFrame.from_dict(history.history).to_csv('clientEHistory_2.csv', index=False)
        return model.get_weights(), train_generator.samples, {}

    def evaluate(self, parameters, config):
        model.set_weights(parameters)
        loss, accuracy, auc, recall, precision = model.evaluate(valid_generator)
        wb = Workbook()
        ws = wb.active
        ws['B2'] = "Model"
        ws['B3'] = "FL_E_2"
        ws['C2'] = "Loss"
        ws['C3'] = loss
        ws['D2'] = "accuracy"
        ws['D3'] = accuracy
        ws['E2'] = "AUC"
        ws['E3'] = auc
        ws['F2'] = "Recall"
        ws['F3'] = recall
        ws['G2'] = "Precision"
        ws['G3'] = precision
        wb.save("FedE_basic_results_2.xlsx")
        return loss, valid_generator.samples, {"accuracy": accuracy, "AUC": auc, 'Recall': recall, 'Precision': precision}


# Start Flower client
fl.client.start_numpy_client(server_address="192.168.1.127:8080", client=CifarClient())

