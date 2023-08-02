import argparse
import os
import numpy as np
import tensorflow as tf
import flwr as fl
from openpyxl import Workbook
import pandas as pd
# import matplotlib.pyplot as plt
from keras_preprocessing.image import ImageDataGenerator

# Make TensorFlow logs less verbose
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"

IMAGE_SHAPE = (224, 224)
TRAINING_DATA_DIR = 'trainF/'
VALID_DATA_DIR = 'valid/'

# EPOCHS = 200  # 1
# BATCH_SIZE = 32


# Define Flower client
class CifarClient(fl.client.NumPyClient):
    def __init__(self, model, train_generator, valid_generator):
        self.model = model
        self.train = train_generator
        self.test = valid_generator

    def get_parameters(self):
        """Get parameters of the local model."""
        raise Exception("Not implemented (server-side parameter initialization)")

    def fit(self, parameters, config):
        """Train parameters on the locally held training set."""

        # Update local model parameters
        self.model.set_weights(parameters)

        # Get hyperparameters for this round
        batch_size: int = config["batch_size"]
        epochs: int = config["local_epochs"]

        # Train the model using hyperparameters from config
        history = self.model.fit(self.train,
                                 steps_per_epoch=self.train.samples // batch_size,
                                 epochs=epochs,
                                 batch_size=batch_size,
                                 validation_data=self.test)

        # Return updated model parameters and results
        pd.DataFrame.from_dict(history.history).to_csv('3clientF_advancedFEDAVGHistory.csv', index=False)
        parameters_prime = self.model.get_weights()
        num_examples_train = len(self.train)
        results = {
            "loss": history.history["loss"][0],
            "accuracy": history.history["accuracy"][0],
            "val_loss": history.history["val_loss"][0],
            "val_accuracy": history.history["val_accuracy"][0],
        }
        return parameters_prime, num_examples_train, results

    def evaluate(self, parameters, config):
        """Evaluate parameters on the locally held test set."""

        # Update local model with global parameters
        self.model.set_weights(parameters)

        # Get config values
        # steps: int = config["val_steps"]

        # Evaluate global model parameters on the local test data and return results
        loss, accuracy, auc, recall, precision = self.model.evaluate(self.test)
        wb = Workbook()
        ws = wb.active
        ws['B2'] = "Model"
        ws['B3'] = "AdvFedF_FEDAVG"
        ws['C2'] = "Loss"
        ws['C3'] = loss
        ws['D2'] = "accuracy"
        ws['D3'] = accuracy
        ws['E2'] = "AUC"
        ws['E3'] = auc
        ws['F2'] = "Recall"
        ws['F3'] = recall
        ws['G2'] = "Precision"
        ws['G3'] = precision
        wb.save("3FedF_advanced_results_FEDAVG.xlsx")
        num_examples_test = len(self.test)
        return loss, num_examples_test, {"accuracy": accuracy}


def main() -> None:
    # Parse command line argument `partition`
    """parser = argparse.ArgumentParser(description="Flower")
    parser.add_argument("--partition", type=int, choices=range(0, 10), required=True)
    args = parser.parse_args()"""

    # Load and compile Keras model

    datagen = tf.keras.preprocessing.image.ImageDataGenerator(
        rescale=1. / 255, validation_split=0.2
    )

    train_generator = datagen.flow_from_directory(
        TRAINING_DATA_DIR,
        shuffle=True,
        target_size=IMAGE_SHAPE,

    )

    valid_generator = datagen.flow_from_directory(
        VALID_DATA_DIR,
        shuffle=False,
        target_size=IMAGE_SHAPE,
    )

    base_model = tf.keras.applications.MobileNetV2(input_shape=(224, 224, 3),
                                                   include_top=False,
                                                   weights='imagenet')

    base_model.trainable = False
    global_average_layer = tf.keras.layers.GlobalAveragePooling2D()
    prediction_layer = tf.keras.layers.Dense(2, activation='softmax')

    model = tf.keras.Sequential([
        base_model,
        global_average_layer,
        prediction_layer
    ])

    base_learning_rate = 0.0001
    model.compile(optimizer=tf.keras.optimizers.RMSprop(learning_rate=base_learning_rate),
                  loss=tf.keras.losses.CategoricalCrossentropy(from_logits=False),
                  metrics=['accuracy', 'AUC', 'Recall', 'Precision'])

    # Start Flower client
    client = CifarClient(model, train_generator, valid_generator)
    fl.client.start_numpy_client("192.168.1.127:8080", client=client)


if __name__ == "__main__":
    main()
